"""
Comparatore Prezzi - Confronta prezzi tra fornitori
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from typing import Optional, List
from datetime import date, datetime, timedelta
from decimal import Decimal
from pydantic import BaseModel

router = APIRouter(prefix="/api/comparatore-prezzi", tags=["Comparatore Prezzi"])


# Placeholder dependencies
async def get_current_user():
    pass

async def get_db():
    pass


# ============================================================================
# MODELS
# ============================================================================

class PriceComparison(BaseModel):
    """Confronto prezzi per un prodotto"""
    prodotto_id: int
    prodotto_nome: str
    categoria: str
    unita_misura: str
    
    # Fornitori
    fornitori: List[dict]  # [{fornitore, prezzo, data_ultimo_ordine}]
    
    # Statistiche
    prezzo_minimo: Decimal
    prezzo_massimo: Decimal
    prezzo_medio: Decimal
    risparmio_potenziale: Decimal
    fornitore_migliore: str


class PriceHistory(BaseModel):
    """Storico prezzi"""
    prodotto_id: int
    fornitore_id: int
    prezzi: List[dict]  # [{data, prezzo}]


class PriceTrend(BaseModel):
    """Trend prezzi"""
    prodotto_id: int
    trend: str  # 'crescente', 'decrescente', 'stabile'
    variazione_percentuale: Decimal
    giorni_analizzati: int


# ============================================================================
# ENDPOINTS
# ============================================================================

@router.get("/", response_model=List[PriceComparison])
async def compare_all_products(
    categoria: Optional[str] = None,
    solo_attivi: bool = True,
    limit: int = Query(100, le=500),
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Confronta prezzi di tutti i prodotti tra fornitori
    
    Ritorna per ogni prodotto:
    - Lista fornitori con prezzi
    - Prezzo min/max/medio
    - Risparmio potenziale
    - Fornitore migliore
    """
    
    query = """
        SELECT 
            p.id as prodotto_id,
            p.nome as prodotto_nome,
            p.categoria,
            p.unita_misura,
            
            -- Prezzi fornitori
            json_agg(
                json_build_object(
                    'fornitore_id', f.id,
                    'fornitore_nome', f.ragione_sociale,
                    'prezzo_unitario', op.prezzo_unitario,
                    'data_ultimo_ordine', MAX(o.data_ordine),
                    'numero_ordini', COUNT(DISTINCT o.id)
                )
            ) as fornitori,
            
            -- Statistiche
            MIN(op.prezzo_unitario) as prezzo_minimo,
            MAX(op.prezzo_unitario) as prezzo_massimo,
            AVG(op.prezzo_unitario) as prezzo_medio,
            MAX(op.prezzo_unitario) - MIN(op.prezzo_unitario) as risparmio_potenziale
            
        FROM prodotti p
        JOIN ordini_prodotti op ON p.id = op.prodotto_id
        JOIN ordini o ON op.ordine_id = o.id
        JOIN fornitori f ON o.fornitore_id = f.id
        WHERE 1=1
    """
    
    params = {}
    
    if categoria:
        query += " AND p.categoria = :categoria"
        params['categoria'] = categoria
    
    if solo_attivi:
        query += " AND p.attivo = true AND f.attivo = true"
    
    query += """
        GROUP BY p.id, p.nome, p.categoria, p.unita_misura
        HAVING COUNT(DISTINCT f.id) >= 2  -- Almeno 2 fornitori
        ORDER BY (MAX(op.prezzo_unitario) - MIN(op.prezzo_unitario)) DESC
        LIMIT :limit
    """
    params['limit'] = limit
    
    results = await db.fetch_all(query, params)
    
    comparisons = []
    
    for row in results:
        # Trova fornitore migliore (prezzo più basso)
        fornitori_list = row['fornitori']
        fornitore_migliore = min(fornitori_list, key=lambda x: x['prezzo_unitario'])
        
        comparisons.append(PriceComparison(
            prodotto_id=row['prodotto_id'],
            prodotto_nome=row['prodotto_nome'],
            categoria=row['categoria'],
            unita_misura=row['unita_misura'],
            fornitori=fornitori_list,
            prezzo_minimo=row['prezzo_minimo'],
            prezzo_massimo=row['prezzo_massimo'],
            prezzo_medio=row['prezzo_medio'],
            risparmio_potenziale=row['risparmio_potenziale'],
            fornitore_migliore=fornitore_migliore['fornitore_nome']
        ))
    
    return comparisons


@router.get("/prodotto/{prodotto_id}")
async def compare_product(
    prodotto_id: int,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Confronto prezzi dettagliato per un singolo prodotto
    
    Include:
    - Tutti i fornitori che vendono il prodotto
    - Storico prezzi ultimi 6 mesi
    - Trend prezzi
    - Raccomandazione fornitore
    """
    
    # Verifica prodotto
    prodotto = await db.fetch_one("""
        SELECT * FROM prodotti WHERE id = :id
    """, {'id': prodotto_id})
    
    if not prodotto:
        raise HTTPException(404, "Prodotto non trovato")
    
    # Prezzi attuali fornitori
    fornitori = await db.fetch_all("""
        SELECT 
            f.id,
            f.ragione_sociale,
            f.email,
            f.telefono,
            op.prezzo_unitario as ultimo_prezzo,
            o.data_ordine as data_ultimo_ordine,
            COUNT(DISTINCT o.id) as numero_ordini_totali,
            AVG(op.prezzo_unitario) as prezzo_medio_storico
        FROM fornitori f
        JOIN ordini o ON f.id = o.fornitore_id
        JOIN ordini_prodotti op ON o.id = op.ordine_id
        WHERE op.prodotto_id = :prodotto_id
        GROUP BY f.id, f.ragione_sociale, f.email, f.telefono, op.prezzo_unitario, o.data_ordine
        HAVING o.data_ordine = (
            SELECT MAX(o2.data_ordine)
            FROM ordini o2
            JOIN ordini_prodotti op2 ON o2.id = op2.ordine_id
            WHERE op2.prodotto_id = :prodotto_id AND o2.fornitore_id = f.id
        )
        ORDER BY op.prezzo_unitario ASC
    """, {'prodotto_id': prodotto_id})
    
    if not fornitori:
        return {
            'prodotto': dict(prodotto),
            'fornitori': [],
            'message': 'Nessun fornitore trovato per questo prodotto'
        }
    
    # Storico prezzi ultimi 6 mesi
    sei_mesi_fa = datetime.now() - timedelta(days=180)
    
    storico = await db.fetch_all("""
        SELECT 
            f.ragione_sociale as fornitore,
            o.data_ordine,
            op.prezzo_unitario
        FROM ordini o
        JOIN ordini_prodotti op ON o.id = op.ordine_id
        JOIN fornitori f ON o.fornitore_id = f.id
        WHERE op.prodotto_id = :prodotto_id
        AND o.data_ordine >= :data_inizio
        ORDER BY o.data_ordine ASC
    """, {'prodotto_id': prodotto_id, 'data_inizio': sei_mesi_fa})
    
    # Calcola trend
    if len(storico) >= 2:
        primo_prezzo = storico[0]['prezzo_unitario']
        ultimo_prezzo = storico[-1]['prezzo_unitario']
        variazione = ((ultimo_prezzo - primo_prezzo) / primo_prezzo) * 100
        
        if abs(variazione) < 5:
            trend = 'stabile'
        elif variazione > 0:
            trend = 'crescente'
        else:
            trend = 'decrescente'
    else:
        variazione = 0
        trend = 'dati_insufficienti'
    
    # Raccomandazione
    fornitore_migliore = fornitori[0]  # Già ordinato per prezzo
    
    raccomandazione = {
        'fornitore_id': fornitore_migliore['id'],
        'fornitore_nome': fornitore_migliore['ragione_sociale'],
        'prezzo': float(fornitore_migliore['ultimo_prezzo']),
        'motivo': []
    }
    
    # Motivi raccomandazione
    if fornitore_migliore['numero_ordini_totali'] >= 5:
        raccomandazione['motivo'].append('Fornitore affidabile (5+ ordini)')
    
    prezzo_medio_mercato = sum(f['ultimo_prezzo'] for f in fornitori) / len(fornitori)
    if fornitore_migliore['ultimo_prezzo'] < prezzo_medio_mercato * Decimal('0.95'):
        raccomandazione['motivo'].append('Prezzo 5%+ sotto media mercato')
    
    if trend == 'decrescente':
        raccomandazione['motivo'].append('Trend prezzi in calo')
    
    return {
        'prodotto': {
            'id': prodotto['id'],
            'nome': prodotto['nome'],
            'categoria': prodotto['categoria'],
            'unita_misura': prodotto['unita_misura']
        },
        'fornitori': [dict(f) for f in fornitori],
        'storico_6_mesi': [dict(s) for s in storico],
        'trend': {
            'tipo': trend,
            'variazione_percentuale': float(variazione),
            'giorni_analizzati': 180
        },
        'raccomandazione': raccomandazione,
        'statistiche': {
            'prezzo_minimo': float(min(f['ultimo_prezzo'] for f in fornitori)),
            'prezzo_massimo': float(max(f['ultimo_prezzo'] for f in fornitori)),
            'prezzo_medio': float(sum(f['ultimo_prezzo'] for f in fornitori) / len(fornitori)),
            'numero_fornitori': len(fornitori)
        }
    }


@router.get("/categoria/{categoria}")
async def compare_by_category(
    categoria: str,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Confronto prezzi per categoria
    
    Mostra prodotti con maggior risparmio potenziale
    """
    
    comparisons = await compare_all_products(
        categoria=categoria,
        solo_attivi=True,
        limit=50,
        current_user=current_user,
        db=db
    )
    
    # Ordina per risparmio potenziale
    comparisons_sorted = sorted(
        comparisons,
        key=lambda x: x.risparmio_potenziale,
        reverse=True
    )
    
    return {
        'categoria': categoria,
        'prodotti_analizzati': len(comparisons_sorted),
        'risparmio_totale_potenziale': sum(c.risparmio_potenziale for c in comparisons_sorted),
        'top_10_risparmi': comparisons_sorted[:10]
    }


@router.get("/fornitori/ranking")
async def rank_suppliers(
    categoria: Optional[str] = None,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Ranking fornitori per convenienza
    
    Ordina fornitori per:
    1. Prezzo medio più basso
    2. Numero prodotti competitivi
    3. Affidabilità (ordini completati)
    """
    
    query = """
        SELECT 
            f.id,
            f.ragione_sociale,
            f.email,
            
            -- Metriche prezzi
            COUNT(DISTINCT op.prodotto_id) as prodotti_venduti,
            AVG(op.prezzo_unitario) as prezzo_medio,
            
            -- Competitività (quante volte è il più economico)
            COUNT(*) FILTER (
                WHERE op.prezzo_unitario = (
                    SELECT MIN(op2.prezzo_unitario)
                    FROM ordini_prodotti op2
                    JOIN ordini o2 ON op2.ordine_id = o2.id
                    WHERE op2.prodotto_id = op.prodotto_id
                )
            ) as volte_piu_economico,
            
            -- Affidabilità
            COUNT(DISTINCT o.id) as ordini_totali,
            COUNT(DISTINCT o.id) FILTER (WHERE o.stato = 'completato') as ordini_completati
            
        FROM fornitori f
        JOIN ordini o ON f.id = o.fornitore_id
        JOIN ordini_prodotti op ON o.id = op.ordine_id
    """
    
    params = {}
    
    if categoria:
        query += """
            JOIN prodotti p ON op.prodotto_id = p.id
            WHERE p.categoria = :categoria
        """
        params['categoria'] = categoria
    else:
        query += " WHERE 1=1"
    
    query += """
        GROUP BY f.id, f.ragione_sociale, f.email
        HAVING COUNT(DISTINCT o.id) >= 3  -- Almeno 3 ordini
        ORDER BY 
            (COUNT(*) FILTER (WHERE op.prezzo_unitario = (
                SELECT MIN(op2.prezzo_unitario)
                FROM ordini_prodotti op2
                WHERE op2.prodotto_id = op.prodotto_id
            ))) DESC,
            AVG(op.prezzo_unitario) ASC
        LIMIT 20
    """
    
    fornitori = await db.fetch_all(query, params)
    
    ranking = []
    
    for idx, f in enumerate(fornitori, 1):
        percentuale_completamento = (
            (f['ordini_completati'] / f['ordini_totali'] * 100)
            if f['ordini_totali'] > 0 else 0
        )
        
        score = (
            (f['volte_piu_economico'] * 10) +  # 10 punti per ogni volta più economico
            (f['prodotti_venduti'] * 2) +       # 2 punti per prodotto
            (percentuale_completamento / 10)     # Max 10 punti per affidabilità
        )
        
        ranking.append({
            'posizione': idx,
            'fornitore_id': f['id'],
            'fornitore': f['ragione_sociale'],
            'email': f['email'],
            'prodotti_venduti': f['prodotti_venduti'],
            'prezzo_medio': float(f['prezzo_medio']),
            'volte_piu_economico': f['volte_piu_economico'],
            'ordini_totali': f['ordini_totali'],
            'ordini_completati': f['ordini_completati'],
            'affidabilita_percentuale': round(percentuale_completamento, 1),
            'score': round(score, 2)
        })
    
    return {
        'categoria': categoria or 'tutte',
        'fornitori': ranking
    }


@router.get("/alert/prezzi-alti")
async def alert_high_prices(
    soglia_percentuale: float = 10.0,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Alert prodotti con prezzi troppo alti
    
    Identifica prodotti dove paghiamo X% in più
    rispetto al fornitore più economico
    """
    
    query = """
        WITH prezzi_correnti AS (
            SELECT 
                p.id as prodotto_id,
                p.nome as prodotto_nome,
                f.id as fornitore_attuale_id,
                f.ragione_sociale as fornitore_attuale,
                op.prezzo_unitario as prezzo_attuale,
                o.data_ordine
            FROM prodotti p
            JOIN ordini_prodotti op ON p.id = op.prodotto_id
            JOIN ordini o ON op.ordine_id = o.id
            JOIN fornitori f ON o.fornitore_id = f.id
            WHERE o.data_ordine = (
                SELECT MAX(o2.data_ordine)
                FROM ordini o2
                JOIN ordini_prodotti op2 ON o2.id = op2.ordine_id
                WHERE op2.prodotto_id = p.id
            )
        ),
        prezzi_minimi AS (
            SELECT 
                op.prodotto_id,
                MIN(op.prezzo_unitario) as prezzo_minimo,
                f.ragione_sociale as fornitore_economico
            FROM ordini_prodotti op
            JOIN ordini o ON op.ordine_id = o.id
            JOIN fornitori f ON o.fornitore_id = f.id
            WHERE o.data_ordine >= NOW() - INTERVAL '6 months'
            GROUP BY op.prodotto_id, f.ragione_sociale
        )
        SELECT 
            pc.prodotto_id,
            pc.prodotto_nome,
            pc.fornitore_attuale,
            pc.prezzo_attuale,
            pm.prezzo_minimo,
            pm.fornitore_economico,
            ((pc.prezzo_attuale - pm.prezzo_minimo) / pm.prezzo_minimo * 100) as percentuale_sovrapprezzo,
            (pc.prezzo_attuale - pm.prezzo_minimo) as risparmio_possibile
        FROM prezzi_correnti pc
        JOIN prezzi_minimi pm ON pc.prodotto_id = pm.prodotto_id
        WHERE ((pc.prezzo_attuale - pm.prezzo_minimo) / pm.prezzo_minimo * 100) >= :soglia
        AND pc.fornitore_attuale_id != pm.fornitore_economico
        ORDER BY percentuale_sovrapprezzo DESC
        LIMIT 50
    """
    
    alerts = await db.fetch_all(query, {'soglia': soglia_percentuale})
    
    risparmio_totale = sum(a['risparmio_possibile'] for a in alerts)
    
    return {
        'soglia_percentuale': soglia_percentuale,
        'prodotti_alert': len(alerts),
        'risparmio_totale_potenziale': float(risparmio_totale),
        'alerts': [dict(a) for a in alerts]
    }


@router.post("/salva-preferenza")
async def save_supplier_preference(
    prodotto_id: int,
    fornitore_id: int,
    motivo: Optional[str] = None,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Salva preferenza fornitore per prodotto
    
    Utile per ricordare scelte anche se non è il più economico
    (es: qualità migliore, consegna più veloce)
    """
    
    await db.execute("""
        INSERT INTO preferenze_fornitori (
            prodotto_id, fornitore_id, motivo, creato_da
        ) VALUES (
            :prod_id, :forn_id, :motivo, :user_id
        )
        ON CONFLICT (prodotto_id) DO UPDATE
        SET fornitore_id = :forn_id,
            motivo = :motivo,
            updated_at = NOW()
    """, {
        'prod_id': prodotto_id,
        'forn_id': fornitore_id,
        'motivo': motivo,
        'user_id': current_user.id
    })
    
    return {
        'success': True,
        'message': 'Preferenza salvata'
    }


@router.get("/export/excel")
async def export_comparison_excel(
    categoria: Optional[str] = None,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Esporta confronto prezzi in Excel
    
    Con:
    - Tabella comparativa
    - Grafici trend
    - Raccomandazioni
    """
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, Reference
    
    # Ottieni dati
    comparisons = await compare_all_products(
        categoria=categoria,
        solo_attivi=True,
        limit=100,
        current_user=current_user,
        db=db
    )
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confronto Prezzi"
    
    # Header
    headers = [
        'Prodotto', 'Categoria', 'Prezzo Min', 'Prezzo Max',
        'Prezzo Medio', 'Risparmio Potenziale', 'Fornitore Migliore'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
    
    # Dati
    for row_idx, comp in enumerate(comparisons, 2):
        ws.cell(row_idx, 1, comp.prodotto_nome)
        ws.cell(row_idx, 2, comp.categoria)
        ws.cell(row_idx, 3, float(comp.prezzo_minimo))
        ws.cell(row_idx, 4, float(comp.prezzo_massimo))
        ws.cell(row_idx, 5, float(comp.prezzo_medio))
        ws.cell(row_idx, 6, float(comp.risparmio_potenziale))
        ws.cell(row_idx, 7, comp.fornitore_migliore)
    
    # Salva
    import os
    output_dir = '/home/claude/azienda-cloud/backend/exports'
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f'confronto_prezzi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    from fastapi.responses import FileResponse
    return FileResponse(
        filepath,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename
    )
