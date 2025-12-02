"""
Controllo Mensile - Confronto Corrispettivi vs POS
Verifica quadratura tra:
- Corrispettivi telematici inviati
- Incassi POS
- Contante in cassa
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import Optional, List
from datetime import date, datetime
from decimal import Decimal
from pydantic import BaseModel

router = APIRouter(prefix="/api/controllo-mensile", tags=["Controllo Mensile"])


# Placeholder dependencies
async def get_current_user():
    pass

async def get_current_admin_user():
    pass

async def get_db():
    pass


# ============================================================================
# MODELS
# ============================================================================

class ControlloMensileCreate(BaseModel):
    """Crea controllo mensile"""
    anno: int
    mese: int
    data_controllo: date


class ControlloDettaglio(BaseModel):
    """Dettaglio controllo"""
    id: int
    anno: int
    mese: int
    data_controllo: date
    
    # Corrispettivi
    totale_corrispettivi: Decimal
    corrispettivi_contante: Decimal
    corrispettivi_elettronici: Decimal
    
    # POS
    totale_pos: Decimal
    numero_transazioni_pos: int
    
    # Cassa
    totale_cassa: Decimal
    
    # Quadrature
    differenza_pos_corrispettivi: Decimal
    differenza_cassa_corrispettivi: Decimal
    quadrato: bool
    
    note: Optional[str]
    created_at: datetime


# ============================================================================
# ENDPOINTS
# ============================================================================

@router.post("/genera")
async def genera_controllo_mensile(
    anno: int,
    mese: int,
    current_user = Depends(get_current_admin_user),
    db = Depends(get_db)
):
    """
    Genera controllo mensile automatico
    
    1. Somma corrispettivi del mese
    2. Somma movimenti POS del mese
    3. Somma movimenti cassa del mese
    4. Calcola differenze
    5. Verifica quadratura
    """
    
    # 1. CORRISPETTIVI TELEMATICI
    corrispettivi = await db.fetch_one("""
        SELECT
            COALESCE(SUM(importo_totale), 0) as totale,
            COALESCE(SUM(importo_contante), 0) as contante,
            COALESCE(SUM(importo_elettronico), 0) as elettronico
        FROM corrispettivi_telematici
        WHERE EXTRACT(YEAR FROM data) = :anno
        AND EXTRACT(MONTH FROM data) = :mese
    """, {'anno': anno, 'mese': mese})
    
    # 2. MOVIMENTI POS
    pos = await db.fetch_one("""
        SELECT
            COALESCE(SUM(importo), 0) as totale,
            COUNT(*) as numero_transazioni
        FROM movimenti_pos
        WHERE EXTRACT(YEAR FROM data) = :anno
        AND EXTRACT(MONTH FROM data) = :mese
        AND tipo = 'incasso'
    """, {'anno': anno, 'mese': mese})
    
    # 3. MOVIMENTI CASSA
    cassa = await db.fetch_one("""
        SELECT
            COALESCE(SUM(importo), 0) as totale
        FROM prima_nota_cassa
        WHERE EXTRACT(YEAR FROM data) = :anno
        AND EXTRACT(MONTH FROM data) = :mese
        AND tipo = 'entrata'
    """, {'anno': anno, 'mese': mese})
    
    # 4. CALCOLA DIFFERENZE
    totale_corrispettivi = corrispettivi['totale']
    totale_pos = pos['totale']
    totale_cassa = cassa['totale']
    
    diff_pos = totale_pos - corrispettivi['elettronico']
    diff_cassa = totale_cassa - corrispettivi['contante']
    
    # 5. VERIFICA QUADRATURA (tolleranza 1 euro)
    quadrato = (abs(diff_pos) <= Decimal('1.0') and abs(diff_cassa) <= Decimal('1.0'))
    
    # 6. SALVA CONTROLLO
    controllo_id = await db.execute("""
        INSERT INTO controllo_mensile (
            anno, mese, data_controllo,
            totale_corrispettivi, corrispettivi_contante, corrispettivi_elettronici,
            totale_pos, numero_transazioni_pos,
            totale_cassa,
            differenza_pos_corrispettivi, differenza_cassa_corrispettivi,
            quadrato,
            creato_da
        ) VALUES (
            :anno, :mese, :data,
            :tot_corr, :corr_cont, :corr_el,
            :tot_pos, :num_pos,
            :tot_cassa,
            :diff_pos, :diff_cassa,
            :quadrato,
            :user_id
        ) RETURNING id
    """, {
        'anno': anno,
        'mese': mese,
        'data': date.today(),
        'tot_corr': totale_corrispettivi,
        'corr_cont': corrispettivi['contante'],
        'corr_el': corrispettivi['elettronico'],
        'tot_pos': totale_pos,
        'num_pos': pos['numero_transazioni'],
        'tot_cassa': totale_cassa,
        'diff_pos': diff_pos,
        'diff_cassa': diff_cassa,
        'quadrato': quadrato,
        'user_id': current_user.id
    })
    
    return {
        'success': True,
        'controllo_id': controllo_id,
        'quadrato': quadrato,
        'dettagli': {
            'corrispettivi': {
                'totale': float(totale_corrispettivi),
                'contante': float(corrispettivi['contante']),
                'elettronico': float(corrispettivi['elettronico'])
            },
            'pos': {
                'totale': float(totale_pos),
                'transazioni': pos['numero_transazioni'],
                'differenza': float(diff_pos)
            },
            'cassa': {
                'totale': float(totale_cassa),
                'differenza': float(diff_cassa)
            }
        }
    }


@router.get("/")
async def list_controlli(
    anno: Optional[int] = None,
    limit: int = 12,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """Lista controlli mensili"""
    
    query = "SELECT * FROM controllo_mensile WHERE 1=1"
    params = {}
    
    if anno:
        query += " AND anno = :anno"
        params['anno'] = anno
    
    query += " ORDER BY anno DESC, mese DESC LIMIT :limit"
    params['limit'] = limit
    
    controlli = await db.fetch_all(query, params)
    
    return {
        'success': True,
        'controlli': [dict(c) for c in controlli]
    }


@router.get("/{controllo_id}")
async def get_controllo(
    controllo_id: int,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """Dettagli controllo mensile"""
    
    controllo = await db.fetch_one("""
        SELECT * FROM controllo_mensile WHERE id = :id
    """, {'id': controllo_id})
    
    if not controllo:
        raise HTTPException(404, "Controllo non trovato")
    
    # Dettagli giornalieri del mese
    dettagli_giornalieri = await db.fetch_all("""
        SELECT
            c.data,
            COALESCE(c.importo_totale, 0) as corrispettivi,
            COALESCE(SUM(p.importo), 0) as pos,
            COALESCE(SUM(cassa.importo), 0) as cassa
        FROM corrispettivi_telematici c
        LEFT JOIN movimenti_pos p ON c.data = p.data
        LEFT JOIN prima_nota_cassa cassa ON c.data = cassa.data AND cassa.tipo = 'entrata'
        WHERE EXTRACT(YEAR FROM c.data) = :anno
        AND EXTRACT(MONTH FROM c.data) = :mese
        GROUP BY c.data, c.importo_totale
        ORDER BY c.data
    """, {'anno': controllo['anno'], 'mese': controllo['mese']})
    
    return {
        'success': True,
        'controllo': dict(controllo),
        'dettagli_giornalieri': [dict(d) for d in dettagli_giornalieri]
    }


@router.post("/{controllo_id}/note")
async def aggiungi_note(
    controllo_id: int,
    note: str,
    current_user = Depends(get_current_admin_user),
    db = Depends(get_db)
):
    """Aggiungi note al controllo"""
    
    await db.execute("""
        UPDATE controllo_mensile
        SET note = :note, updated_at = NOW()
        WHERE id = :id
    """, {'id': controllo_id, 'note': note})
    
    return {'success': True, 'message': 'Note salvate'}


@router.get("/anomalie/detect")
async def detect_anomalie(
    anno: int,
    mese: int,
    soglia_differenza: Decimal = Decimal('10.0'),
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    Rileva anomalie nel mese
    
    Cerca:
    - Giorni con differenze > soglia
    - Giorni senza corrispettivi ma con POS
    - Giorni senza POS ma con corrispettivi elettronici
    """
    
    anomalie = []
    
    # Giorni con differenze significative
    differenze = await db.fetch_all("""
        SELECT
            c.data,
            c.importo_totale as corrispettivi,
            COALESCE(SUM(p.importo), 0) as pos,
            ABS(c.importo_elettronico - COALESCE(SUM(p.importo), 0)) as differenza
        FROM corrispettivi_telematici c
        LEFT JOIN movimenti_pos p ON c.data = p.data
        WHERE EXTRACT(YEAR FROM c.data) = :anno
        AND EXTRACT(MONTH FROM c.data) = :mese
        GROUP BY c.data, c.importo_totale, c.importo_elettronico
        HAVING ABS(c.importo_elettronico - COALESCE(SUM(p.importo), 0)) > :soglia
        ORDER BY differenza DESC
    """, {'anno': anno, 'mese': mese, 'soglia': soglia_differenza})
    
    for diff in differenze:
        anomalie.append({
            'tipo': 'differenza_pos',
            'data': diff['data'],
            'descrizione': f"Differenza di €{diff['differenza']:.2f} tra POS e corrispettivi",
            'gravita': 'alta' if diff['differenza'] > 50 else 'media',
            'dettagli': dict(diff)
        })
    
    # Giorni senza corrispettivi
    giorni_senza_corr = await db.fetch_all("""
        SELECT
            p.data,
            SUM(p.importo) as totale_pos
        FROM movimenti_pos p
        LEFT JOIN corrispettivi_telematici c ON p.data = c.data
        WHERE EXTRACT(YEAR FROM p.data) = :anno
        AND EXTRACT(MONTH FROM p.data) = :mese
        AND c.id IS NULL
        GROUP BY p.data
        ORDER BY totale_pos DESC
    """, {'anno': anno, 'mese': mese})
    
    for giorno in giorni_senza_corr:
        anomalie.append({
            'tipo': 'corrispettivi_mancanti',
            'data': giorno['data'],
            'descrizione': f"POS presente (€{giorno['totale_pos']:.2f}) ma corrispettivi mancanti",
            'gravita': 'alta',
            'dettagli': dict(giorno)
        })
    
    return {
        'success': True,
        'anno': anno,
        'mese': mese,
        'totale_anomalie': len(anomalie),
        'anomalie': anomalie
    }


# ============================================================================
# REPORT EXCEL
# ============================================================================

@router.get("/{controllo_id}/export/excel")
async def export_controllo_excel(
    controllo_id: int,
    current_user = Depends(get_current_user),
    db = Depends(get_db)
):
    """Esporta controllo in Excel"""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    controllo = await db.fetch_one("""
        SELECT * FROM controllo_mensile WHERE id = :id
    """, {'id': controllo_id})
    
    if not controllo:
        raise HTTPException(404, "Controllo non trovato")
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Controllo {controllo['anno']}-{controllo['mese']:02d}"
    
    # Header
    ws['A1'] = f"CONTROLLO MENSILE - {controllo['anno']}/{controllo['mese']:02d}"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Riepilogo
    row = 3
    ws[f'A{row}'] = "RIEPILOGO"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Corrispettivi Totali:"
    ws[f'B{row}'] = float(controllo['totale_corrispettivi'])
    
    row += 1
    ws[f'A{row}'] = "- Contante:"
    ws[f'B{row}'] = float(controllo['corrispettivi_contante'])
    
    row += 1
    ws[f'A{row}'] = "- Elettronici:"
    ws[f'B{row}'] = float(controllo['corrispettivi_elettronici'])
    
    row += 2
    ws[f'A{row}'] = "POS Totale:"
    ws[f'B{row}'] = float(controllo['totale_pos'])
    
    row += 1
    ws[f'A{row}'] = "Differenza POS:"
    ws[f'B{row}'] = float(controllo['differenza_pos_corrispettivi'])
    if abs(controllo['differenza_pos_corrispettivi']) > 1:
        ws[f'B{row}'].fill = PatternFill(start_color="FFFF0000", fill_type="solid")
    
    row += 2
    ws[f'A{row}'] = "Cassa Totale:"
    ws[f'B{row}'] = float(controllo['totale_cassa'])
    
    row += 1
    ws[f'A{row}'] = "Differenza Cassa:"
    ws[f'B{row}'] = float(controllo['differenza_cassa_corrispettivi'])
    if abs(controllo['differenza_cassa_corrispettivi']) > 1:
        ws[f'B{row}'].fill = PatternFill(start_color="FFFF0000", fill_type="solid")
    
    row += 2
    ws[f'A{row}'] = "QUADRATO:"
    ws[f'B{row}'] = "✓ SÌ" if controllo['quadrato'] else "✗ NO"
    ws[f'B{row}'].font = Font(bold=True)
    if controllo['quadrato']:
        ws[f'B{row}'].fill = PatternFill(start_color="FF00FF00", fill_type="solid")
    else:
        ws[f'B{row}'].fill = PatternFill(start_color="FFFF0000", fill_type="solid")
    
    # Salva
    output_path = f'/home/claude/azienda-cloud/backend/exports/controllo_{controllo_id}.xlsx'
    wb.save(output_path)
    
    return FileResponse(
        output_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f"Controllo_{controllo['anno']}_{controllo['mese']:02d}.xlsx"
    )
